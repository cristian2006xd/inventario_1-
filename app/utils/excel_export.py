"""Exportación del inventario a Excel (.xlsx) con los encabezados oficiales
en español, en el mismo orden que la tabla real, para que el archivo se
pueda reabrir/reimportar sin fricción.
"""
import io
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.utils.campos import CAMPOS_EN_ORDEN_OFICIAL


def exportar_inventario_excel(bienes: list) -> io.BytesIO:
    columnas = [c["label"] for c in CAMPOS_EN_ORDEN_OFICIAL]
    filas = []
    for bien in bienes:
        fila = {}
        for campo in CAMPOS_EN_ORDEN_OFICIAL:
            valor = getattr(bien, campo["attr"])
            if hasattr(valor, "isoformat"):
                valor = valor.isoformat()
            elif valor is not None and not isinstance(valor, (str, int, float, bool)):
                valor = float(valor)
            fila[campo["label"]] = valor
        filas.append(fila)

    df = pd.DataFrame(filas, columns=columnas)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Inventario")
    buffer.seek(0)
    return buffer


def nombre_archivo_exportacion(termino_busqueda: str | None) -> str:
    marca_tiempo = datetime.now().strftime("%Y%m%d_%H%M%S")
    if termino_busqueda:
        return f"inventario_filtrado_{marca_tiempo}.xlsx"
    return f"inventario_completo_{marca_tiempo}.xlsx"


def generar_plantilla_excel() -> io.BytesIO:
    """Genera un .xlsx vacío con los encabezados oficiales del inventario,
    listo para llenar y volver a subir en Inventario Previo -> Cargar Excel.

    Los encabezados son exactamente los que reconoce el importador (ver
    `app/utils/excel_import.py` y los `aliases` en `app/utils/campos.py`),
    así que un archivo basado en esta plantilla se mapea sin fricción.
    """
    columnas = [c["label"] for c in CAMPOS_EN_ORDEN_OFICIAL]
    df = pd.DataFrame(columns=columnas)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Plantilla")
        libro = writer.book
        hoja = writer.sheets["Plantilla"]

        atributos_requeridos = {c["attr"] for c in CAMPOS_EN_ORDEN_OFICIAL if c.get("requerido")}
        relleno_normal = PatternFill("solid", fgColor="1D4ED8")
        relleno_requerido = PatternFill("solid", fgColor="B45309")
        fuente_encabezado = Font(color="FFFFFF", bold=True)

        for indice, campo in enumerate(CAMPOS_EN_ORDEN_OFICIAL, start=1):
            celda = hoja.cell(row=1, column=indice)
            celda.font = fuente_encabezado
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celda.fill = relleno_requerido if campo["attr"] in atributos_requeridos else relleno_normal
            ancho = max(14, min(34, len(campo["label"]) + 4))
            hoja.column_dimensions[get_column_letter(indice)].width = ancho

        hoja.row_dimensions[1].height = 32
        hoja.freeze_panes = "A2"

        hoja_instrucciones = libro.create_sheet("Instrucciones")
        instrucciones = [
            "Cómo llenar esta plantilla",
            "",
            "- Cada columna corresponde a un campo oficial del inventario de INAMHI.",
            "- Las columnas en NARANJA (Código del Bien, (BLD) o (BCA), Bien, "
            "Serie/ Identificación, Ubicación de Bodega, Custodio Actual, Estado Bien) "
            "son obligatorias.",
            "- El 'Código del Bien' debe ser único. Si dejas la celda vacía, el sistema "
            "le asignará un código provisional automáticamente al momento de subir el archivo.",
            "- No cambies los nombres de los encabezados de la primera fila: el sistema "
            "los usa para reconocer cada columna.",
            "- Puedes dejar en blanco las columnas que no apliquen a un bien en particular.",
            "- Una vez lleno, sube este archivo desde Inventario Previo -> Cargar Excel.",
        ]
        for fila, texto in enumerate(instrucciones, start=1):
            celda = hoja_instrucciones.cell(row=fila, column=1, value=texto)
            if fila == 1:
                celda.font = Font(bold=True, size=13)
        hoja_instrucciones.column_dimensions["A"].width = 100

    buffer.seek(0)
    return buffer
